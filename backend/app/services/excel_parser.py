"""Parser for Screener.in `.xlsx` exports.

Source of truth = the `Data Sheet` tab. Sections are located by their header
labels in column A (never by hardcoded row numbers, since rows drift across
companies / Screener versions). See PLANNING §3.
"""
from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl import load_workbook


class NotAScreenerFile(Exception):
    """Raised when an uploaded workbook is not a recognizable Screener export."""


DATA_SHEET = "Data Sheet"

# Section header label (as it appears in col A) -> internal section key.
SECTION_HEADERS: dict[str, str] = {
    "META": "meta",
    "PROFIT & LOSS": "pnl",
    "QUARTERS": "quarters",
    "BALANCE SHEET": "balance_sheet",
    "CASH FLOW:": "cash_flow",
    "PRICE:": "price",
    "DERIVED:": "derived",
}


@dataclass
class Section:
    key: str
    periods: list[str] = field(default_factory=list)
    # line item label -> list of values aligned to `periods`
    items: dict[str, list[Optional[float]]] = field(default_factory=dict)


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if s in ("", "-", "—", "NA", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _fy_label(value: Any) -> Optional[str]:
    """Convert a Report Date cell into an FY label, e.g. 'FY2023'.

    Screener dates are typically end-of-period (Mar of the FY)."""
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        dt = value
    elif isinstance(value, _dt.date):
        dt = _dt.datetime(value.year, value.month, value.day)
    else:
        s = str(value).strip()
        # try a few common string formats
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%b-%y", "%b %Y", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = _dt.datetime.strptime(s, fmt)
                break
            except ValueError:
                dt = None  # type: ignore
        else:
            # maybe already an FY-ish string
            m = re.search(r"(20\d{2}|19\d{2})", s)
            if m:
                return f"FY{m.group(1)}"
            return s or None
        if dt is None:
            return s or None
    # Screener fiscal year ends in March; treat Jan-Mar as the FY of that year,
    # Apr-Dec as the next FY. Screener already aligns to FY-end, so just use year.
    fy = dt.year
    return f"FY{fy}"


class ScreenerWorkbook:
    """Parsed representation of a Screener Data Sheet."""

    def __init__(self, rows: list[list[Any]]):
        self._rows = rows
        self.sections: dict[str, Section] = {}
        self.meta: dict[str, Any] = {}
        self.company_name: str = ""
        self._parse()

    # ---- section / period detection ----
    def _find_report_date_row(self, start: int) -> Optional[int]:
        """From `start`, find the next row whose col A begins with 'Report Date'."""
        for i in range(start, min(start + 6, len(self._rows))):
            a = _norm(self._rows[i][0]) if self._rows[i] else ""
            if a.lower().startswith("report date"):
                return i
        return None

    def _periods_from_row(self, row: list[Any]) -> tuple[list[str], list[int]]:
        """Return (fy_labels, column_indices) from a Report Date row (cols B..)."""
        labels: list[str] = []
        cols: list[int] = []
        for ci in range(1, len(row)):
            label = _fy_label(row[ci])
            if label:
                labels.append(label)
                cols.append(ci)
        return labels, cols

    def _parse(self) -> None:
        rows = self._rows
        n = len(rows)

        # locate section header rows
        header_rows: list[tuple[int, str]] = []
        for i, row in enumerate(rows):
            if not row:
                continue
            a = _norm(row[0]).upper().rstrip()
            for header_label, key in SECTION_HEADERS.items():
                if a == header_label or a == header_label.rstrip(":"):
                    header_rows.append((i, key))
                    break

        if not any(k in {k2 for _, k2 in header_rows} for k in ("pnl",)):
            raise NotAScreenerFile(
                "Missing 'PROFIT & LOSS' section — not a recognizable Screener Data Sheet."
            )

        header_idx_set = [hi for hi, _ in header_rows]

        for pos, (hrow, key) in enumerate(header_rows):
            end = header_rows[pos + 1][0] if pos + 1 < len(header_rows) else n

            if key == "meta":
                self._parse_meta(hrow + 1, end)
                continue

            rd = self._find_report_date_row(hrow + 1)
            if rd is None or rd >= end:
                # sections like DERIVED may not have a report-date row; capture as plain items
                section = Section(key=key)
                self._parse_items(section, hrow + 1, end, cols=None, periods=[])
                self.sections[key] = section
                continue

            labels, cols = self._periods_from_row(rows[rd])
            section = Section(key=key, periods=labels)
            self._parse_items(section, rd + 1, end, cols=cols, periods=labels)
            self.sections[key] = section

    def _parse_meta(self, start: int, end: int) -> None:
        for i in range(start, end):
            row = self._rows[i]
            if not row:
                continue
            label = _norm(row[0])
            if not label:
                continue
            # value is the first non-empty cell after col A
            val: Any = None
            for ci in range(1, len(row)):
                if _norm(row[ci]) != "":
                    val = row[ci]
                    break
            key = label.strip()
            if key.upper() == "COMPANY NAME":
                self.company_name = _norm(val)
                self.meta["company_name"] = self.company_name
            else:
                fv = _to_float(val)
                self.meta[key] = fv if fv is not None else _norm(val)

    def _parse_items(
        self,
        section: Section,
        start: int,
        end: int,
        cols: Optional[list[int]],
        periods: list[str],
    ) -> None:
        for i in range(start, end):
            row = self._rows[i]
            if not row:
                continue
            label = _norm(row[0])
            if not label:
                continue
            if label.upper() in SECTION_HEADERS:
                break
            if cols is None:
                # single-value style section
                val = None
                for ci in range(1, len(row)):
                    if _norm(row[ci]) != "":
                        val = _to_float(row[ci])
                        break
                section.items[label] = [val]
            else:
                values = [_to_float(row[ci]) if ci < len(row) else None for ci in cols]
                section.items[label] = values

    # ---- accessors ----
    def to_dict(self) -> dict[str, Any]:
        return {
            "company_name": self.company_name,
            "meta": self.meta,
            "sections": {
                k: {"periods": s.periods, "items": s.items}
                for k, s in self.sections.items()
            },
        }


def _read_rows(path: str) -> list[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if DATA_SHEET not in wb.sheetnames:
            # case-insensitive lookup
            match = next(
                (s for s in wb.sheetnames if s.strip().lower() == DATA_SHEET.lower()),
                None,
            )
            if match is None:
                raise NotAScreenerFile(
                    f"Workbook has no '{DATA_SHEET}' tab. Tabs found: {wb.sheetnames}"
                )
            ws = wb[match]
        else:
            ws = wb[DATA_SHEET]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows
    finally:
        wb.close()


def parse_screener_file(path: str) -> ScreenerWorkbook:
    """Parse a Screener .xlsx at `path`. Raises NotAScreenerFile if invalid."""
    try:
        rows = _read_rows(path)
    except NotAScreenerFile:
        raise
    except Exception as exc:  # corrupt / not an xlsx
        raise NotAScreenerFile(f"Could not read workbook: {exc}") from exc
    return ScreenerWorkbook(rows)
